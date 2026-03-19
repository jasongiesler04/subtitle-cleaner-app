import streamlit as st
import pandas as pd
import re
import os

# --- Settings ---
#st.set_page_config(layout="wide")

# --- Page title ---
st.title("Script Processor")

# -------------------------------------------------
# Regex patterns and manual character blocks
# -------------------------------------------------
CHARACTER_REGEX = r"([A-Z][^\s:]*?(?:\s+[A-Za-z0-9'’\-&]+)*):"

STOPWORDS = {
    "because", "remember", "what", "where", "when", "why", "how",
    "this", "that", "these", "those", "if", "then", "but", "and", "or",
    "um", "uh", "erm", "hmm", "so", "now", "we", "oh", "i", "i'm", "it",
    "said"
}

BLOCKED_CHARACTERS = {
    "whispers", "sings", "sing", "remember",
    "raps", "impersonating", "distorted"
}

SENTENCE_SPLIT_REGEX = r'(?:\n|(?<=[.!?])\s+)'

SENTENCE_REGEX = re.compile(r".*?[.!?](?:['\"\)\]]*)(?=\s|$)")


# -------------------------------------------------
# Helper functions
# -------------------------------------------------

# Function to try and identify genuine character names
def looks_like_character(name):
    if not isinstance(name, str):
        return False

    # Remove punctuation so "Um...update" becomes "um update"
    cleaned = re.sub(r"[^\w\s]", " ", name.lower())
    words = cleaned.split()

    # Reject if any stopword appears
    if any(w in STOPWORDS for w in words):
        return False

    # Reject if more than 3 words
    if len(words) > 3:
        return False

    return True

# Function to split and clean the subtitle text
def preprocess_subtitles(df):
    df = df.copy()
    df["orig_row_id"] = df.index

    # Split on newlines
    df["subtitle_text_split"] = df["Subtitle Text"].str.split("\n")
    df_new = df.explode("subtitle_text_split")

    # Split into sentence-like fragments
    df_new["subtitle_text_split"] = df_new["subtitle_text_split"].str.split(SENTENCE_SPLIT_REGEX)
    df_new = df_new.explode("subtitle_text_split")

    # Mark last fragment of original subtitle row
    df_new["is_last_fragment"] = (
        df_new.groupby("orig_row_id").cumcount(ascending=False) == 0
    )

    # Clean fragments
    df_new["subtitle_text_split"] = df_new["subtitle_text_split"].str.strip()
    df_new = df_new[
        df_new["subtitle_text_split"].notna() &
        df_new["subtitle_text_split"].ne("")
    ]

    return df_new

# Function to extract all raw characters from subtitle text
def extract_all_characters(df):
    exploded = (
        df["subtitle_text_split"]
        .str.extractall(CHARACTER_REGEX)[0]
        .dropna()
        .str.replace(":", "", regex=False)
        .str.strip()
        .unique()
    )
    return list(exploded)

# Function to separate out potential characters and non-characters
def auto_classify_characters(characters):
    chars, nonchars = [], []
    for sp in characters:
        if looks_like_character(sp) and sp.lower() not in BLOCKED_CHARACTERS:
            chars.append(sp)
        else:
            nonchars.append(sp)
    return chars, nonchars


# -------------------------------------------------
# Step 2: Upload + Character Classification
# -------------------------------------------------

if "step" not in st.session_state:
    st.session_state.step = 1

st.subheader("Step 1: Upload Raw Excel Subtitle File")
uploaded_file = st.file_uploader(" ", type=["xlsx"])

if uploaded_file and st.session_state.step == 1:

    df_raw = pd.read_excel(uploaded_file)
    st.session_state.df_raw = df_raw

    # Clean subtitle text
    df_clean = preprocess_subtitles(df_raw)

    # Extract characters
    all_characters = extract_all_characters(df_clean)
    likely_chars, likely_nonchars = auto_classify_characters(all_characters)

    # Initialise session lists on first load
    if "char_status" not in st.session_state:
        # True = character, False = non-character
        st.session_state.char_status = {
            sp: (sp in likely_chars) for sp in all_characters
        }
        # green first, then red
        st.session_state.sorted_speakers = (
                sorted(likely_chars) + sorted(likely_nonchars)
        )

    st.subheader("Step 2: Confirm Characters")
    st.write("Click a name to toggle between Character (🟢) and Non‑Character (🔴).")

    # Multi-column layout (3 columns)
    num_cols = 3
    cols = st.columns(num_cols)

    # Render buttons in fixed order
    for i, sp in enumerate(st.session_state.sorted_speakers):
        is_char = st.session_state.char_status[sp]
        icon = "🟢" if is_char else "🔴"
        label = f"{icon} {sp}"

        col = cols[i % num_cols]

        with col:
            if st.button(label, key=f"toggle_{sp}", use_container_width=True):
                st.session_state.char_status[sp] = not is_char
                st.rerun()

    # Continue button
    st.markdown("<br>", unsafe_allow_html=True)
    if st.button("**Continue to Process Subtitles**"):
        # Store characters in the SAME order as Step 1
        st.session_state.final_characters = [
            sp for sp in st.session_state.sorted_speakers
            if st.session_state.char_status[sp]
        ]

        # Also store the full ordered list for display
        st.session_state.final_order = st.session_state.sorted_speakers

        st.session_state.step = 2
        st.rerun()


# -------------------------------------------------
# Step 3: Subtitle Processing
# -------------------------------------------------

if uploaded_file and st.session_state.step == 2:

    # -------------------------------------------------
    # Show selected characters in the SAME layout as the above step
    # -------------------------------------------------
    st.subheader("Step 2: Confirmed Characters")

    num_cols = 3
    cols = st.columns(num_cols)

    for i, sp in enumerate(st.session_state.final_order):
        is_char = sp in st.session_state.final_characters
        icon = "🟢" if is_char else "🔴"
        label = f"{icon} {sp}"

        col = cols[i % num_cols]
        with col:
            st.markdown(f"**{label}**")

    df = st.session_state.df_raw
    clean_characters = st.session_state.final_characters

    st.markdown("<br>", unsafe_allow_html=True)
    st.subheader("Step 3: Subtitle Converter")
    st.write("Using your confirmed character list.")

    # -------------------------------
    # 1. Split lines, clean text and extract characters
    # -------------------------------

    # Clean subtitle text
    df_new = preprocess_subtitles(df)

    # Extract character
    df_new["Character"] = df_new["subtitle_text_split"].str.extract(CHARACTER_REGEX)

    # Clean extracted character names
    df_new["Character"] = (
        df_new["Character"]
        .str.replace(":", "", regex=False)
        .str.strip()
    )

    # Keep only characters from user selection
    df_new["Character"] = df_new["Character"].where(
        df_new["Character"].isin(clean_characters),
        None
    )

    # Forward-fill speakers
    df_new["Character"] = df_new["Character"].ffill()

    # Remove character prefix from text
    def remove_real_character_prefix(text):
        for sp in clean_characters:
            if text.startswith(sp + ":"):
                return text[len(sp) + 1:].strip()
        return text

    df_new["line_text"] = (
        df_new["subtitle_text_split"]
        .str.replace("_x000D_", "", regex=False)
        .str.replace('"', "", regex=False)
        .str.replace("\r", "", regex=False)
        .str.strip()
        .apply(remove_real_character_prefix)
    )

    # -------------------------------
    # 2. Sentence reconstruction
    # -------------------------------
    def rebuild_block(block):
        rows = []
        buffer = ""
        sentence_start_time = None
        character_written = False

        for _, row in block.iterrows():
            text = row["line_text"]
            if not text:
                continue

            if buffer == "":
                sentence_start_time = row["Start Time"]
                sentence_end_time = row["End Time"]
                last_end_time = row["End Time"]

            buffer += (" " if buffer else "") + text

            while True:
                match = SENTENCE_REGEX.match(buffer)
                if not match:
                    break

                sentence = match.group().strip()
                if not character_written:
                    sentence = f"{row['Character']}: {sentence}"
                    character_written = True

                row_fully_consumed = match.group().endswith(text) and row["is_last_fragment"]
                end_time = row["End Time"] if row_fully_consumed else sentence_end_time

                rows.append({
                    "Character": row["Character"],
                    "Subtitle Text": sentence,
                    "Start Time": sentence_start_time,
                    "End Time": end_time,
                })

                buffer = buffer[len(match.group()):].strip()
                sentence_start_time = row["Start Time"]

        if buffer:
            sentence = buffer.strip()
            if not character_written:
                sentence = f"{block.iloc[0]['Character']}: {sentence}"

            rows.append({
                "Character": block.iloc[0]["Character"],
                "Subtitle Text": sentence,
                "Start Time": sentence_start_time,
                "End Time": last_end_time,
            })

        return pd.DataFrame(rows)

    df_final = (
        df_new
        .groupby(df_new['Character'].ne(df_new['Character'].shift()).cumsum(), group_keys=False)
        .apply(rebuild_block)
        .reset_index(drop=True)
    )

    # -------------------------------
    # 3. Merge consecutive same-character & timestamp rows
    # -------------------------------
    df_final['merge_group'] = (
        (df_final['Character'] != df_final['Character'].shift()) |
        (df_final['Start Time'] != df_final['Start Time'].shift())
    ).cumsum()

    df_final = (
        df_final
        .groupby('merge_group', as_index=False, sort=False)
        .agg({
            'Character': 'first',
            'Subtitle Text': lambda x: ' '.join(x),
            'Start Time': 'first',
            'End Time': 'last',
        })
    ).drop(columns='merge_group')

    # -------------------------------
    # 4. Duration
    # -------------------------------
    def time_to_ms(ts):
        h, m, rest = ts.split(":")
        s, ms = rest.split(",")
        return int(h)*3600000 + int(m)*60000 + int(s)*1000 + int(ms)

    def ms_to_time(ms):
        h = ms // 3600000
        m = (ms % 3600000) // 60000
        s = (ms % 60000) // 1000
        ms = ms % 1000
        return f"{h:02d}:{m:02d}:{s:02d},{ms:03d}"

    df_final["Duration (ms)"] = (
        df_final["End Time"].apply(time_to_ms)
        - df_final["Start Time"].apply(time_to_ms)
    )

    df_final["Duration"] = df_final["Duration (ms)"].apply(ms_to_time)
    df_final = df_final.drop(columns="Duration (ms)")

    # -------------------------------
    # 5. Add row count and word count
    # -------------------------------
    df_final["Index"] = range(1, len(df_final) + 1)
    df_final["Word Count"] = ""

    df_final = df_final[
        ["Index", "Character", "Subtitle Text", "Start Time", "End Time", "Duration", "Word Count"]
    ]

    st.success("Subtitles processed successfully!")

    # Store for downloads
    st.session_state.df_final = df_final
    st.session_state.base_name = os.path.splitext(uploaded_file.name)[0]


    # -------------------------------
    # 6. Prepare Excel download
    # -------------------------------
    from io import BytesIO
    from openpyxl import load_workbook

    buffer = BytesIO()

    # write dataframe
    df_final.to_excel(buffer, index=False, engine="openpyxl")
    buffer.seek(0)

    # open workbook from memory
    wb = load_workbook(buffer)
    ws = wb.active

    wb.calculation.fullCalcOnLoad = True

    # insert formulas
    for row in range(2, len(df_final) + 2):
        ws[f"G{row}"].value = (
            f'=IF(TRIM(C{row})="",0,'
            f'LEN(TRIM(C{row}))-LEN(SUBSTITUTE(TRIM(C{row})," ",""))+1'
            f'-IF(B{row}=B{row - 1},0,'
            f'LEN(TRIM(B{row}))-LEN(SUBSTITUTE(TRIM(B{row})," ",""))+1)'
            f')'
        )

    # save back to memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)


    # -------------------------------
    # 7. Prepare SRT download
    # -------------------------------
    def format_srt_time(ts):
        if isinstance(ts, str):
            ts = ts.replace('.', ',')
            if ',' not in ts:
                ts += ',000'
            return ts
        else:
            total_ms = int(ts * 1000)
            hours = total_ms // 3600000
            minutes = (total_ms % 3600000) // 60000
            seconds = (total_ms % 60000) // 1000
            milliseconds = total_ms % 1000
            return f"{hours:02d}:{minutes:02d}:{seconds:02d},{milliseconds:03d}"


    print(df_final.columns.tolist())
    print(next(df_final.itertuples()))

    srt_lines = []
    for i, row in enumerate(df_final.itertuples(), 1):
        start = format_srt_time(row._4)  # Start Time
        end = format_srt_time(row._5)  # End Time
        text = row._3  # Subtitle Text
        srt_lines.append(f"{i}\n{start} --> {end}\n{text}\n\n")

    srt_file_name = f"{st.session_state.base_name}.srt"
    with open(srt_file_name, "w", encoding="utf-8") as f:
        f.writelines(srt_lines)

    with open(srt_file_name, "rb") as f:
        srt_data = f.read()

    # -------------------------------
    # 8. Download Buttons
    # -------------------------------

    st.subheader("Step 4: Download Processed Subtitles")

    col1, col2 = st.columns(2)

    with col1:
        st.download_button(
            label="Download Excel",
            data=output,
            file_name=f"{st.session_state.base_name}-processed.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

    with col2:
        st.download_button(
            label="Download SRT",
            data=srt_data,
            file_name=srt_file_name,
            mime="text/plain",
            use_container_width=True
        )
